from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


REKAP_HEADERS = [
    "Test Case ID",
    "Module",
    "Scenario",
    "Task Turunan",
    "Expected Result",
    "Actual Result",
    "Status",
    "Tester",
    "Execution Date",
    "Evidence Path",
    "Notes",
]

DEFECT_HEADERS = [
    "Defect ID",
    "Test Case ID",
    "Module",
    "Summary",
    "Steps to Reproduce",
    "Expected Result",
    "Actual Result",
    "Severity",
    "Status",
    "Evidence Path",
    "Owner",
    "Notes",
]

EXPECTED_BY_ID = {
    "BBGR-01": "Assessment berhasil tersimpan dan report siswa menampilkan hasil sesuai input",
    "BBGR-02": "Sistem menolak penyimpanan dan menampilkan pesan validasi",
    "BBGR-03": "Sistem dinyatakan Fail karena output tidak sesuai siswa/input yang dipilih",
    "BBAE-01": "Bobot berhasil dihitung, ternormalisasi, dan dapat digunakan dalam assessment",
    "BBAE-02": "Sistem menolak proses dan menampilkan validasi",
    "BBAE-03": "Sistem dinyatakan Fail karena bobot tidak layak digunakan",
    "BBPJ-01": "Kriteria berhasil tersimpan dan muncul pada rubrik assessment",
    "BBPJ-02": "Sistem menolak penyimpanan dan menampilkan validasi",
    "BBPJ-03": "Sistem dinyatakan Fail jika perubahan berdampak ke data yang tidak semestinya",
    "BBAD-01": "Akun berhasil dibuat dan user hanya melihat fitur sesuai role",
    "BBAD-02": "Sistem menolak penyimpanan dan menampilkan validasi",
    "BBAD-03": "Sistem dinyatakan Fail karena akses tidak sesuai role/status akun",
}

TASKS_BY_ID = {
    "BBAD-01": "Login admin; buat user baru; pilih role Wali Kelas/Evaluator; set akses kelas/mapel; login sebagai user baru; verifikasi dashboard dan Student Management sesuai role.",
    "BBAD-02": "Login admin; kirim data user tidak lengkap atau tidak valid; verifikasi backend/UI menolak dan menampilkan validasi.",
    "BBAD-03": "Login admin; buat user guru dan user nonaktif; login sebagai guru; akses halaman admin; verifikasi akses ditolak; verifikasi user nonaktif tidak bisa login.",
    "BBAE-01": "Login ATL Expert; isi TFN valid; lengkapi pairwise comparison; proses Fuzzy-AHP; verifikasi bobot ternormalisasi; simpan bobot.",
    "BBAE-02": "Login ATL Expert; isi TFN tidak valid atau pairwise belum lengkap; verifikasi sistem menolak proses/simpan.",
    "BBAE-03": "Login ATL Expert; coba simpan bobot kosong/tidak valid/tidak ternormalisasi; verifikasi sistem mencegah hasil tidak layak.",
    "BBGR-01": "Login guru; pilih kelas; pilih siswa; pilih topik; isi seluruh rubrik ATL; simpan assessment; buka report; verifikasi hasil siswa muncul.",
    "BBGR-02": "Login guru; pilih konteks assessment; kosongkan siswa atau indikator wajib; klik simpan; verifikasi validasi tampil.",
    "BBGR-03": "Login guru; simpan assessment untuk Student A; buka report Student B; verifikasi data Student A tidak tertukar dan skor sesuai input.",
    "BBPJ-01": "Login PJ Mapel; buka Criteria Management; tambah kriteria/subskill/indikator valid; simpan; verifikasi muncul pada konteks/rubrik.",
    "BBPJ-02": "Login PJ Mapel; coba simpan kriteria kosong atau kode duplikat; verifikasi sistem menolak dan menampilkan validasi.",
    "BBPJ-03": "Login PJ Mapel; ubah kriteria pada satu konteks; verifikasi perubahan tidak merusak konteks lain atau data assessment lama.",
}


def ensure_workbook(path: Path, headers: list[str]):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row == 0 or [cell.value for cell in ws[1]][: len(headers)] != headers:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    ws.freeze_panes = "A2"
    for index, _header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = 24
    return wb, ws


def clear_data_rows(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def iter_specs(suite):
    for spec in suite.get("specs", []):
        yield spec, suite
    for child in suite.get("suites", []):
        yield from iter_specs(child)


def module_from_title(title: str) -> str:
    if title.startswith("BBGR"):
        return "Guru/Evaluator - Input Assessment ATL"
    if title.startswith("BBAE"):
        return "ATL Expert - Pembobotan Fuzzy-AHP"
    if title.startswith("BBPJ"):
        return "PJ Matkul/PJ Mapel - Manajemen Kriteria"
    if title.startswith("BBAD"):
        return "Admin Akademik - Role dan Akses"
    return "Blackbox E2E"


def clean_error(results) -> str:
    for result in results or []:
        errors = result.get("errors") or []
        if errors:
            message = errors[0].get("message") or errors[0].get("value") or ""
            return re.sub(r"\x1b\[[0-9;]*m", "", str(message)).strip()
        error = result.get("error")
        if error:
            return re.sub(r"\x1b\[[0-9;]*m", "", str(error.get("message") or error)).strip()
    return ""


def evidence_path(playwright_report_dir: Path, test_id: str) -> str:
    if playwright_report_dir.exists():
        return str(playwright_report_dir)
    return f"Document Output/playwright-report/ ({test_id})"


def main():
    parser = argparse.ArgumentParser(description="Update rekap dan defect log dari Playwright JSON report.")
    parser.add_argument("--json", default="Document Output/terminal-output/playwright-results.json")
    parser.add_argument("--output", default="Document Output")
    parser.add_argument("--tester", default="")
    args = parser.parse_args()

    root = Path.cwd()
    json_path = root / args.json
    output_dir = root / args.output
    report_dir = output_dir / "playwright-report"
    rekap_path = output_dir / "rekap-blackbox-testing.xlsx"
    defect_path = output_dir / "defect-log.xlsx"

    data = json.loads(json_path.read_text(encoding="utf-8"))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    rekap_wb, rekap_ws = ensure_workbook(rekap_path, REKAP_HEADERS)
    defect_wb, defect_ws = ensure_workbook(defect_path, DEFECT_HEADERS)
    clear_data_rows(rekap_ws)
    clear_data_rows(defect_ws)

    defect_index = 1
    for suite in data.get("suites", []):
        for spec, _suite in iter_specs(suite):
            title = spec.get("title") or ""
            match = re.match(r"^(BB(?:GR|AE|PJ|AD)-\d{2})\s*-\s*(.*)$", title)
            test_id = match.group(1) if match else title.split(" - ", 1)[0]
            scenario = title
            module = module_from_title(test_id)
            tests = spec.get("tests") or []
            results = [result for test in tests for result in test.get("results", [])]
            status = "Pass" if all(test.get("status") == "expected" for test in tests) else "Fail"
            error_text = clean_error(results)
            actual = "Sesuai expected result." if status == "Pass" else (error_text[:300] or "Test gagal. Lihat Playwright report.")
            evidence = evidence_path(report_dir, test_id)

            rekap_ws.append([
                test_id,
                module,
                scenario,
                TASKS_BY_ID.get(test_id, ""),
                EXPECTED_BY_ID.get(test_id, ""),
                actual,
                status,
                args.tester,
                now,
                evidence,
                "",
            ])

            if status == "Fail":
                defect_ws.append([
                    f"DEF-{defect_index:03d}",
                    test_id,
                    module,
                    f"{test_id} gagal saat eksekusi Playwright",
                    f"Jalankan test case {test_id} melalui npm run test:e2e.",
                    EXPECTED_BY_ID.get(test_id, ""),
                    actual,
                    "High",
                    "Open",
                    evidence,
                    "",
                    "Lengkapi detail setelah review screenshot/video/trace.",
                ])
                defect_index += 1

    rekap_wb.save(rekap_path)
    defect_wb.save(defect_path)
    print(f"Updated: {rekap_path}")
    print(f"Updated: {defect_path}")


if __name__ == "__main__":
    main()
