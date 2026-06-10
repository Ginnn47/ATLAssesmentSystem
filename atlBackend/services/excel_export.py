from io import BytesIO
from re import sub

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def safe_excel_filename(value):
    cleaned = sub(r"[^A-Za-z0-9._-]+", "_", str(value or "ATL_Report")).strip("_")
    cleaned = cleaned or "ATL_Report"
    return cleaned if cleaned.lower().endswith(".xlsx") else f"{cleaned}.xlsx"


def build_report_workbook(meta, columns, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ATL Score List"
    numeric_keys = {"no", "score", "thinking", "research", "communication", "social", "selfManagement"}

    title = "ATL SCORE LIST"
    generated_at = meta.get("generatedAt") or "-"
    identity_rows = [
        ("CLASS", meta.get("className") or "-"),
        ("SUBJECT", meta.get("subject") or "-"),
        ("SUB TOPIC", meta.get("subTopic") or "-"),
        ("TOTAL STUDENTS", meta.get("rowCount") or len(rows)),
        ("EXPORT TIME", generated_at),
    ]

    headers = [column.get("label", "").upper() for column in columns]
    keys = [column.get("key") for column in columns]
    last_column = max(1, len(headers))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(row=1, column=1, value=title)

    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    sheet.cell(row=1, column=1).font = Font(bold=True, size=16, color="111827")

    label_fill = PatternFill("solid", fgColor="FEF3C7")
    for row_index, (label, value) in enumerate(identity_rows, start=3):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        label_cell.fill = label_fill
        label_cell.font = Font(bold=True, color="92400E")
        value_cell.font = Font(bold=True, color="111827")
        label_cell.alignment = Alignment(horizontal="left")
        value_cell.alignment = Alignment(horizontal="left")

    header_row = 10
    header_fill = PatternFill("solid", fgColor="F9B208")
    thin_side = Side(style="thin", color="D6D3D1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_offset, row in enumerate(rows, start=1):
        excel_row = header_row + row_offset
        for column_index, key in enumerate(keys, start=1):
            value = row.get(key, "-")
            if key in numeric_keys and isinstance(value, str):
                try:
                    value = float(value)
                except ValueError:
                    pass
            cell = sheet.cell(row=excel_row, column=column_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if key in numeric_keys and isinstance(value, (int, float)):
                cell.number_format = "0.00" if key != "no" else "0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_column)}{max(header_row, header_row + len(rows))}"

    for column_index in range(1, last_column + 1):
        letter = get_column_letter(column_index)
        max_length = len(str(sheet.cell(row=header_row, column=column_index).value or ""))
        for cell in sheet[letter]:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[letter].width = min(max(max_length + 3, 12), 30)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
